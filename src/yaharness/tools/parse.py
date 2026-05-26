"""File-parsing tools for GAIA-style problems.

Each tool reads a path (scoped via the caller's own filesystem discipline —
these are pure parsers, they do not enforce scope) and returns a textual
representation suitable for LLM consumption. Output is capped to keep
context tractable.

Tools
-----
- ``ParseCsvTool``    — CSV via stdlib ``csv``
- ``ParseExcelTool``  — XLSX via ``openpyxl``
- ``ParsePdfTool``    — PDF text extraction via ``pypdf``
- ``ParseEmlTool``    — RFC 822 email via stdlib ``email``
"""

from __future__ import annotations

import csv
from email import policy
from email.parser import BytesParser
from pathlib import Path
from typing import Any, ClassVar

from . import ToolResult

_MAX_CELL_CHARS = 200


def _truncate_cell(value: object) -> str:
    s = "" if value is None else str(value)
    if len(s) > _MAX_CELL_CHARS:
        return s[:_MAX_CELL_CHARS] + "...[truncated]"
    return s


def _render_table(rows: list[list[str]]) -> str:
    """Pipe-separated rendering — compact and easy for LLMs to read."""
    return "\n".join(" | ".join(cell for cell in row) for row in rows)


def _parse_csv(path: Path, max_rows: int) -> ToolResult:
    rows: list[list[str]] = []
    with path.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        for i, row in enumerate(reader):
            if i >= max_rows + 1:  # +1 to include header row
                break
            rows.append([_truncate_cell(c) for c in row])
    if not rows:
        return ToolResult(ok=True, output="", metadata={"rows": 0, "columns": 0})
    return ToolResult(
        ok=True,
        output=_render_table(rows),
        metadata={"rows": len(rows) - 1, "columns": len(rows[0])},
    )


def _parse_excel(path: Path, sheet: str | None, max_rows: int) -> ToolResult:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet is not None else wb[wb.sheetnames[0]]
        rows: list[list[str]] = []
        for i, raw in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows + 1:
                break
            rows.append([_truncate_cell(c) for c in raw])
    finally:
        wb.close()
    return ToolResult(
        ok=True,
        output=_render_table(rows),
        metadata={
            "rows": max(0, len(rows) - 1),
            "columns": len(rows[0]) if rows else 0,
            "sheet": sheet or wb.sheetnames[0],
            "sheets": list(wb.sheetnames),
        },
    )


def _parse_pdf(path: Path, max_pages: int, max_chars: int) -> ToolResult:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    n_pages = len(reader.pages)
    chunks: list[str] = []
    for i, page in enumerate(reader.pages):
        if i >= max_pages:
            break
        text = page.extract_text() or ""
        chunks.append(f"--- page {i + 1} ---\n{text.strip()}")
    output = "\n\n".join(chunks)
    truncated = False
    if len(output) > max_chars:
        output = output[:max_chars] + "\n...[truncated]"
        truncated = True
    return ToolResult(
        ok=True,
        output=output,
        metadata={
            "pages": n_pages,
            "pages_read": min(max_pages, n_pages),
            "truncated": truncated,
        },
    )


def _parse_eml(path: Path, max_chars: int) -> ToolResult:
    with path.open("rb") as fh:
        msg = BytesParser(policy=policy.default).parse(fh)
    headers = {
        "From": str(msg.get("From", "")),
        "To": str(msg.get("To", "")),
        "Cc": str(msg.get("Cc", "")),
        "Subject": str(msg.get("Subject", "")),
        "Date": str(msg.get("Date", "")),
    }
    body_part = msg.get_body(preferencelist=("plain", "html"))
    body_text = ""
    body_type = ""
    if body_part is not None:
        body_type = str(body_part.get_content_type())
        try:
            content = body_part.get_content()
            body_text = str(content) if not isinstance(content, str) else content
        except (LookupError, ValueError):
            body_text = ""
    header_block = "\n".join(f"{k}: {v}" for k, v in headers.items() if v)
    body_block = body_text.strip()
    output = f"{header_block}\n\n{body_block}".strip()
    truncated = False
    if len(output) > max_chars:
        output = output[:max_chars] + "\n...[truncated]"
        truncated = True
    return ToolResult(
        ok=True,
        output=output,
        metadata={
            "headers": headers,
            "body_content_type": body_type,
            "truncated": truncated,
        },
    )


class ParseCsvTool:
    name = "parse_csv"
    description = "Parse a CSV file; returns header + first N rows as a pipe-separated table."
    parameters_schema: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "max_rows": {"type": "integer", "default": 100},
        },
        "required": ["path"],
    }

    async def execute(self, **kwargs: Any) -> ToolResult:
        path = kwargs.get("path")
        max_rows = kwargs.get("max_rows", 100)
        if not isinstance(path, str):
            return ToolResult(ok=False, output="", error="`path` must be a string")
        if not isinstance(max_rows, int) or max_rows <= 0:
            return ToolResult(ok=False, output="", error="`max_rows` must be a positive int")
        p = Path(path)
        if not p.exists() or not p.is_file():
            return ToolResult(ok=False, output="", error=f"no such file: {path}")
        try:
            return _parse_csv(p, max_rows)
        except (OSError, csv.Error, UnicodeDecodeError) as exc:
            return ToolResult(ok=False, output="", error=f"csv parse error: {exc}")


class ParseExcelTool:
    name = "parse_excel"
    description = (
        "Parse an XLSX file; returns first N rows of the chosen sheet (default first sheet)."
    )
    parameters_schema: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string"},
            "max_rows": {"type": "integer", "default": 100},
        },
        "required": ["path"],
    }

    async def execute(self, **kwargs: Any) -> ToolResult:
        path = kwargs.get("path")
        sheet = kwargs.get("sheet")
        max_rows = kwargs.get("max_rows", 100)
        if not isinstance(path, str):
            return ToolResult(ok=False, output="", error="`path` must be a string")
        if sheet is not None and not isinstance(sheet, str):
            return ToolResult(ok=False, output="", error="`sheet` must be a string or omitted")
        if not isinstance(max_rows, int) or max_rows <= 0:
            return ToolResult(ok=False, output="", error="`max_rows` must be a positive int")
        p = Path(path)
        if not p.exists() or not p.is_file():
            return ToolResult(ok=False, output="", error=f"no such file: {path}")
        try:
            return _parse_excel(p, sheet, max_rows)
        except KeyError as exc:
            return ToolResult(ok=False, output="", error=f"no such sheet: {exc}")
        except (OSError, ValueError) as exc:
            return ToolResult(ok=False, output="", error=f"excel parse error: {exc}")


class ParsePdfTool:
    name = "parse_pdf"
    description = "Extract text from a PDF; returns up to N pages of text."
    parameters_schema: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "max_pages": {"type": "integer", "default": 20},
            "max_chars": {"type": "integer", "default": 50_000},
        },
        "required": ["path"],
    }

    async def execute(self, **kwargs: Any) -> ToolResult:
        path = kwargs.get("path")
        max_pages = kwargs.get("max_pages", 20)
        max_chars = kwargs.get("max_chars", 50_000)
        if not isinstance(path, str):
            return ToolResult(ok=False, output="", error="`path` must be a string")
        if not isinstance(max_pages, int) or max_pages <= 0:
            return ToolResult(ok=False, output="", error="`max_pages` must be a positive int")
        if not isinstance(max_chars, int) or max_chars <= 0:
            return ToolResult(ok=False, output="", error="`max_chars` must be a positive int")
        p = Path(path)
        if not p.exists() or not p.is_file():
            return ToolResult(ok=False, output="", error=f"no such file: {path}")
        try:
            return _parse_pdf(p, max_pages, max_chars)
        except Exception as exc:  # pypdf raises a varied exception zoo
            return ToolResult(ok=False, output="", error=f"pdf parse error: {exc}")


class ParseEmlTool:
    name = "parse_eml"
    description = "Parse an RFC 822 .eml email; returns headers + plaintext body."
    parameters_schema: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "max_chars": {"type": "integer", "default": 50_000},
        },
        "required": ["path"],
    }

    async def execute(self, **kwargs: Any) -> ToolResult:
        path = kwargs.get("path")
        max_chars = kwargs.get("max_chars", 50_000)
        if not isinstance(path, str):
            return ToolResult(ok=False, output="", error="`path` must be a string")
        if not isinstance(max_chars, int) or max_chars <= 0:
            return ToolResult(ok=False, output="", error="`max_chars` must be a positive int")
        p = Path(path)
        if not p.exists() or not p.is_file():
            return ToolResult(ok=False, output="", error=f"no such file: {path}")
        try:
            return _parse_eml(p, max_chars)
        except (OSError, ValueError) as exc:
            return ToolResult(ok=False, output="", error=f"eml parse error: {exc}")


__all__ = ["ParseCsvTool", "ParseEmlTool", "ParseExcelTool", "ParsePdfTool"]
